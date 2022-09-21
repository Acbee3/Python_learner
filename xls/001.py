from openpyxl import load_workbook, Workbook

wd = Workbook()
wd.active.iter_cols()
wd.active['A']
wd.active['1']
wd.save('./new_xlsx.xlsx')